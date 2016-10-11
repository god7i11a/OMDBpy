""">>> import requests
>>> r = requests.get('https://github.com/timeline.json')
>>> r.json()


    r = requests.get('http://www.omdbapi.com/?t=Batman&page=1&r=xml')
    print dir(r)
    print r.text


long term solution: use csv for proper file diffs and convert to xlsx
at end


import csv
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

f = open('file.csv')
reader = csv.reader(f, delimiter=':')
for row in reader:
    ws.append(row)
f.close()

wb.save('file.xlsx')

"""

from requests import get as rget
from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd
import csv
 
_VERBOSE=False
_DOIT=True

keyL = [u'Title', u'Year', u'Series / Episode / ID', u'B-R',
        u'Runtime', u'DLed', u'Director', u'Actors', u'tomatoMeter',
        u'imdbRating', u'Plot', u'tomatoConsensus', u'Genre',
        u'Website', u'Awards', u'Language', u'Country', u'BoxOffice', u'Type']

searchL = [ "Title","Year","imdbID","Type" ]

TYPE='movie'   # default search mode

def csv_from_excel(fN):
    
    wb = xlrd.open_workbook(fN+'.xlsx')
    sh = wb.sheet_by_name('Sheet1')
    csv_file = open(fN+'.csv', 'wb')
    wr = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
    
    for rownum in xrange(sh.nrows):
        wr.writerow(sh.row_values(rownum))
    
    csv_file.close()

def make_not_found(movieN):
    theD = {field: ''  for field in keyL}
    theD[u'Title']= '%s'%movieN
    return theD

def _searchByName(name):
    r=rget('http://www.omdbapi.com/?s=%s'%name)
    return r.json()['Search']

def _getData(movieN, yr, seid, type):
    # need to add in epise=, season=, y= etc
    reqStr = '?r=json&tomatoes=true&type=%s'%(type or TYPE.lower())
    if yr: reqStr = reqStr+'&year=%s'%yr
    #seid='%s'%seid
    if not seid:
        reqStr=reqStr+'&t=%s'%movieN
    elif seid.startswith('tt'):
        reqStr = reqStr+'&i=%s'%seid
    else:
        # parse the string
        x = seid.split('E')
        theS = x[0][1:]
        if len(x) == 1:
            theE = None
        elif len(x)==2:
            theE = x[1]

        reqStr=reqStr+'&t=%s'%movieN
        if theS: reqStr = reqStr+'&season=%s'%theS
        if theE: reqStr = reqStr+'&episode=%s'%theE
            
    if _VERBOSE: print reqStr
    if _DOIT:
        r = rget('http://www.omdbapi.com/'+reqStr)
        res=r.json()
    else:
        res = make_not_found(movieN)
        res['Response']=True
    return res

def _splitName(movieN, part=1):
    left = movieN.find('[')
    if left==-1:
        leftN = movieN
        rightN = None
    else:
        leftN, rightN = movieN.split('[')
        rightN=rightN.strip()[:-1]

    if part==0: return leftN
    if part==1: return rightN
    if part==2: return (leftN, rightN)

def _joinNames(leftN, rightN):
    name= '%s '%leftN
    if rightN: name=name+' [%s]'%rightN
    return name

def _getSearchName(movieN):
    searchN = _splitName(movieN) or movieN
    return _normalizeSearchN(searchN)

def _normalizeSearchN(movieN):
    if type(movieN) is type(3):
        movieN = '%s'%movieN
    elif movieN.lower().endswith(' a'):
         movieN='A '+movieN[:-1]
    elif movieN.lower().endswith(' the'):
        movieN='The '+movieN[:-3]
    movieN=movieN.strip()
    if movieN[-1]==',': movieN=movieN[:-1]
    return movieN

def _alphabetizeTitle(titleN):
    movieN,searchN = _splitName(titleN, part=2)
    movieN=movieN.strip()
    if movieN[-1]==',': movieN=movieN[:-1]
    if movieN.startswith('A '):
        movieN=movieN[2:]+', A'
    elif movieN.startswith('The '):
        movieN=movieN[4:]+', The'
    return _joinNames( movieN, searchN )

def getData(movieL, badFP):
    
    movieN, yr, seid, br, run, dled = movieL[0:6]
    _type = movieL[-1]

    if dled=='x':                   # we have the data already, doanatouch
        res = movieL
        if False:
            res[0]=_alphabetizeTitle( movieN )
            
    elif movieN.startswith('%%'):   # placeholder
        res=make_not_found(movieN)
        # switch modes
        TYPE=movieN[2:]
    else:                           # we need to get data
        searchN = _getSearchName(movieN)
        print 'getting info for %s ...'%searchN
        res=_getData(searchN, yr, seid, _type)
        # check for {"Response":"False","Error":"Movie not found!"}
        if res['Response'] == "False":
            print 'not found!!!'
            if badFP: badFP.write(movieN.encode('utf8')+'\n')
            res=make_not_found(movieN)
            res['DLed']=''
        else:
            res['DLed']='x'

        # TODO: fix the following so we use newfound name as appropriate 
        res[u'Title']=_alphabetizeTitle( movieN )
        res[u'Series / Episode / ID'] = seid
        res[u'B-R'] = br

    if type(res) is type({}):
        res=[res[key]  for key in keyL] 

    return res

def needs_ID():
    wb = load_workbook(filename=infoXLS, read_only=True)
    ws = wb['AllInfo'] # ws is now an IterableWorksheet

    newwb = Workbook()
    newws=newwb.active
    newws.title='AllInfo'
    
    for row in ws.rows:
        dl = row[5].value
        if dl is None:
            newws.append([ row[i].value for i in range(len(keyL))] )

    newwb.save(idXLS)

def fillID():
    wb = load_workbook(filename=idXLS, read_only=True)
    ws = wb['AllInfo'] # ws is now an IterableWorksheet

    newwb = Workbook()
    newws=newwb.active
    newws.title='AllInfo'
    
    for row in ws.rows:
        movieL = [ row[i].value for i in range(len(keyL))]
        res = getData(movieL, None)
        newws.append(res)

    newwb.save(fillidXLS)

def getDiscL(save):
    wb = load_workbook(filename=infoXLS, read_only=False)
    ws = wb['AllInfo'] # ws is now an IterableWorksheet

    titleL=[]
    for row in ws.rows:
        valL = [ row[i].value for i in range(len(keyL))]
        titleL.append( valL )
    
    if save: wb.save(backXLS)
    return titleL

def getAll(movieN):
    wb = Workbook()
    ws=wb.active
    ws.title='AllInfo'
    ws.append(keyL)
    badFP = open(badF, 'w')
    
    movieL=_searchByName(movieN)
    for movieD in movieL:
        name = movieD['Title']
        imdbID = movieD['imdbID']
        typ=movieD['Type']
        yr = movieD['Year']
        resD = _getData(name, yr, imdbID, typ)
        resD[u'Series / Episode / ID'] = imdbID
        resD['B-R']='???'
        resD['DLed']='x'
        if _VERBOSE:
            keys =  (u'Title', u'Year', u'Series / Episode / ID', u'Runtime', u'Director', u'Actors',
                     u'tomatoMeter', u'imdbRating', u'Plot',
                     u'Genre', u'Type')
            print [resD[key].encode('utf-8') for key in keys]
            print
        ws.append([resD[key] for key in keyL])
    wb.save('%s.xlsx'%movieN)

def parseNYT():
    fp = open('1000best.html', 'r')
    data = fp.readlines()
    fp.close()

    wb = Workbook()
    ws=wb.active
    ws.title='AllInfo'
    ws.append(keyL)
    badFP = open(badF, 'w')    

    titleL = []
    for line in data:
        if not line.startswith('<td><a href='): continue
        start = line.find('w\">')
        name=line[start+3:-11]
        leftP = name.find('(')
        movieN = name[0:leftP]
        date = name[leftP+1:-1]
        print movieN, date
        resD = _getData(movieN, date, None, None)

        if resD['Response'] == "False":
            print 'not found!!!'
            if badFP:
                try:
                    badFP.write(movieN.encode('utf8')+'\n')
                except UnicodeDecodeError:
                    badFP.write(movieN+'\n')
            res=make_not_found(movieN)
            resD['DLed']=''
        else:
            resD['DLed']='x'
            try:
                resD[u'Series / Episode / ID']=resD['imdbID']
            except KeyError:
                resD[u'Series / Episode / ID']=''
                
            resD['B-R']=''
            ws.append([resD[key] for key in keyL])
    wb.save('nytimes1000best.xlsx')
    
def main(save=False):
    
    titleL = getDiscL(save)
    
    wb = Workbook()
    ws=wb.active
    ws.title='AllInfo'
    ws.append(keyL)
    badFP = open(badF, 'w')
    
    for title in titleL[1:]:
        res=getData(title, ws, badFP)
        ws.append(res)

    if save: wb.save(infoXLS)
    badFP.close()


if __name__ == '__main__':

    OMDBdir = '/home/godzilla/Desktop/OMDBpy'
    idXLS = '%s/dvd-needs-info.xlsx'%OMDBdir            # need info list if dl / col5 is None in dvd-info
    fillidXLS = '%s/dvd-new-info.xlsx'%OMDBdir          # new info 
    infoXLS= '%s/dvd-info.xlsx'%OMDBdir                 # entire list
    backXLS = '%s/dvd-info-bak.xlsx'%OMDBdir
    badF = '%s/bad-movie-names.txt'%OMDBdir

    import argparse

    parser = argparse.ArgumentParser(description='OMDB API processor')
    parser.add_argument('-r', '--run', action='store_true', default=False,     help='Run the whole list %s and refill'%infoXLS)
    parser.add_argument('-c', '--convert', action='store', default=None,     help='convert filename to csv or vice versa')    
    parser.add_argument('-i', '--need', action='store_true', default=False,
                        help='Generate new need info list %s from whole list %s'% (idXLS, infoXLS) )
    parser.add_argument('-f', '--fill', action='store_true', default=False,
                        help='Get new info from id\'s in %s and put it into %s'% (idXLS, fillidXLS) )
    parser.add_argument('-s', '--search', action='store_true', default=False,     help='Search for name')
    parser.add_argument('-v', '--verbose', action='store_true', default=False,     help='Verbose')
    parser.add_argument('-n', '--name', default=False,     help='movie name')
    parser.add_argument('-t', '--nytimes',   action='store_true', default=False,     help='Run NY Times listing')
    args = parser.parse_args()

    _VERBOSE = args.verbose
    
    if args.run:     main(save=True) # fill the entire dvd-info list with new information
    if args.need:    needs_ID()    # parses dvd-info for no DL 'x' and puts them into dvd-needs-id list
    if args.fill:    fillID()      # parses dvd-needs-id list and fills dvd-new-info list
    if args.search:  getAll(args.name)
    if args.nytimes: parseNYT()
    if args.convert: csv_from_excel(args.convert)
